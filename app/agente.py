import os
import re
import json
import base64
import logging
import pathlib
import httpx
from openai import AsyncOpenAI
from app.db import cargar_historial, guardar_mensajes
from app.tools import TOOLS_SCHEMA, tool_buscar_producto, tool_ver_stock, tool_generar_cotizacion
from app.motor import (
    buscar_por_codigo,
    buscar_por_codigo_prefijo,
    buscar_por_tipo_medida_marca,
    buscar_texto_libre,
    formatear_resultado,
    formatear_multi_marca,
    formatear_lista,
    extraer_cantidad,
    extraer_descuento,
    df as motor_df,
    IGV,
    _stock_total,
    _aliases_marcas,
    MEDIDA_NOMINAL,
)

logger = logging.getLogger(__name__)

_client = AsyncOpenAI(api_key=os.getenv("OPENAI_API_KEY"))
_MODEL  = "gpt-4.1-mini"

# Corrección OCR: "codo" + tipo SAE en la misma línea → "casco" (los codos no llevan subtipo SAE)
_RE_CODO    = re.compile(r'\bcodo\b', re.IGNORECASE)
_RE_SAE     = re.compile(r'\bR\d{1,2}(T/M|T)?\b', re.IGNORECASE)


def _corregir_codo_ocr(texto: str) -> str:
    lineas = []
    for linea in texto.splitlines():
        if _RE_CODO.search(linea) and _RE_SAE.search(linea):
            linea = _RE_CODO.sub('casco', linea)
        lineas.append(linea)
    return '\n'.join(lineas)


def _enriquecer_tipo_ferrula(parsed_list: list[dict]) -> list[dict]:
    """Si el parser devolvió tipo='FERRULA' sin subtipo SAE pero la línea original
    contiene R1/R2/R12/etc., inyecta el subtipo para que el motor filtre correctamente."""
    for item in parsed_list:
        if item.get("tipo", "").upper() == "FERRULA":
            m = _RE_SAE.search(item.get("linea_original", ""))
            if m:
                item["tipo"] = f"FERRULA {m.group(0).upper()}"
    return parsed_list


# Conjunto de medidas hidráulicas conocidas (valores de MEDIDA_NOMINAL)
_MEDIDAS_CONOCIDAS    = set(MEDIDA_NOMINAL.values())
_RE_FRACCION_MEDIDA   = re.compile(r'\b\d+\s+\d+/\d+\b|\b\d+/\d+\b')


def _extraer_fracciones_medida(texto: str) -> list[str]:
    """Extrae fracciones de pulgada conocidas del texto, en orden de aparición."""
    return [m.strip() for m in _RE_FRACCION_MEDIDA.findall(texto)
            if m.strip() in _MEDIDAS_CONOCIDAS]


def _corregir_medidas_ocr(parsed_list: list[dict]) -> list[dict]:
    """Corrección determinista: si GPT redondeó medidas, las restaura desde linea_original.
    Compara posición a posición las fracciones del texto original vs las emitidas por el parser."""
    for item in parsed_list:
        fracs_orig = _extraer_fracciones_medida(item.get("linea_original", ""))
        if not fracs_orig:
            continue

        medidas_gpt = [m.strip() for m in (item.get("medidas") or []) if m.strip()]
        medida_gpt  = (item.get("medida") or "").strip()

        if medidas_gpt:
            corregidas = list(medidas_gpt)
            for i, gpt_val in enumerate(medidas_gpt):
                if i < len(fracs_orig) and gpt_val != fracs_orig[i] and gpt_val in _MEDIDAS_CONOCIDAS:
                    corregidas[i] = fracs_orig[i]
            if corregidas != medidas_gpt:
                item["medidas"] = corregidas
                item["medida"] = ""
                logger.info(f"_corregir_medidas_ocr: {medidas_gpt} → {corregidas}")
        elif medida_gpt and medida_gpt in _MEDIDAS_CONOCIDAS:
            if fracs_orig[0] != medida_gpt:
                logger.info(f"_corregir_medidas_ocr: medida {medida_gpt!r} → {fracs_orig[0]!r}")
                item["medida"] = fracs_orig[0]

    return parsed_list


# Preguntas de intención sobre marcas → GPT directo (antes de cualquier búsqueda)
_INTENT_MARCAS = re.compile(
    r'\b(qu[eé]\s+marcas?|en\s+qu[eé]\s+marcas?|qu[eé]\s+marca\s+tiene|'
    r'qu[eé]\s+marca\s+tienen|en\s+cu[aá]les?\s+marcas?)\b', re.IGNORECASE
)

# Fichas técnicas almacenadas en app/data/fichas/{prefijo_3dig}.docx
_FICHAS_DIR   = pathlib.Path(__file__).parent / "data" / "fichas"
_INTENT_FICHA = re.compile(r'\bficha\b', re.IGNORECASE)
_MIME_DOCX    = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"

# Inteligencia comercial: demandas no encontradas en catálogo
_DEMANDAS_FILE = pathlib.Path(__file__).parent / "data" / "demandas_no_catalogo.jsonl"


def _registrar_demanda(tipo: str, medida: str, cantidad: int, linea_original: str, numero_wa: str, motivo: str = "tamaño") -> None:
    """Registra un ítem no encontrado en catálogo para análisis comercial.
    motivo: 'tamaño' = grupo existe pero no esa medida | 'producto' = grupo no existe."""
    import datetime
    entrada = {
        "fecha":          datetime.datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "motivo":         motivo,
        "tipo":           tipo or "",
        "medida":         medida or "",
        "cantidad":       cantidad,
        "linea_original": linea_original,
        "numero_wa":      numero_wa[-4:],
    }
    try:
        with open(_DEMANDAS_FILE, "a", encoding="utf-8") as f:
            f.write(json.dumps(entrada, ensure_ascii=False) + "\n")
    except Exception as e:
        logger.warning(f"_registrar_demanda: no se pudo escribir: {e}")


def _tipo_existe_en_catalogo(tipo: str) -> bool:
    """Verifica si el tipo/grupo de producto existe en catálogo (sin filtrar por medida)."""
    if not tipo:
        return False
    return not buscar_por_tipo_medida_marca(tipo=tipo).empty


async def _enviar_ficha_wa(numero_wa: str, grupo: str) -> bool:
    """Sube el .docx al Media API de WhatsApp y lo envía como documento.
    Retorna True si se envió, False si el archivo no existe."""
    archivo = _FICHAS_DIR / f"{grupo}.docx"
    if not archivo.exists():
        return False

    token    = os.getenv("WHATSAPP_TOKEN")
    phone_id = os.getenv("PHONE_NUMBER_ID")
    headers  = {"Authorization": f"Bearer {token}"}

    async with httpx.AsyncClient() as client:
        with open(archivo, "rb") as f:
            r_upload = await client.post(
                f"https://graph.facebook.com/v19.0/{phone_id}/media",
                headers=headers,
                data={"messaging_product": "whatsapp", "type": _MIME_DOCX},
                files={"file": (archivo.name, f, _MIME_DOCX)},
                timeout=30,
            )
        r_upload.raise_for_status()
        media_id = r_upload.json()["id"]

        await client.post(
            f"https://graph.facebook.com/v19.0/{phone_id}/messages",
            headers={**headers, "Content-Type": "application/json"},
            json={
                "messaging_product": "whatsapp",
                "to": numero_wa,
                "type": "document",
                "document": {
                    "id": media_id,
                    "filename": f"ficha_tecnica_{grupo}.docx",
                },
            },
            timeout=10,
        )
    return True

_PARSER_PROMPT = """\
Eres un parser de consultas comerciales para CISGE Perú.
Extrae del texto estos campos y devuelve SOLO JSON válido sin texto adicional:
{
  "subfamilias": [],
  "tipo": "",
  "medida": "",
  "medidas": [],
  "marca": "",
  "color": "",
  "presion": "",
  "angulo": "",
  "cola": "",
  "doble_hex": false,
  "ferrula_tm": "",
  "es_saludo": false
}
subfamilias: lista de subfamilias ERP posibles: "MANGUERAS HIDRAULICAS", "ESPIGAS I", "ESPIGAS II", "FERRULAS", "ADAPTADORES I", "ADAPTADORES II", "VALVULAS", etc.
tipo: tipo de producto exacto del catálogo. Si el usuario especificó subtipo (R1/R2/R12/etc.), inclúyelo (ej: "ESPIGA MACHO NPT R2"). Si no especificó, usa el prefijo base (ej: "ESPIGA MACHO NPT").
medida: primera medida — fracción de pulgada (1/2, 3/4, 1, 1 1/4) o mm para silicona/industrial: extrae SOLO el número sin "mm" (ej: 19mm → "19", 38mm → "38", 25mm → "25")
medidas: si el usuario menciona dos medidas separadas por "x" (ej: "3/8 x 3/8", "1/4 x 1/2"), extrae la lista ordenada: ["3/8", "3/8"]. Si hay una sola medida, dejar vacío [].
marca: marca oficial: QF, JDEFLEX, VITILLO, MACTUBI, AF, LT, DME, etc.
color: A=Amarillo, N=Negro, R=Rojo — solo si se menciona explícitamente
presion: si se menciona presión de trabajo
angulo: ángulo de la conexión — "45" si dice 45°/45 grados, "90" si dice 90°/90 grados, "" si no especifica (recta por defecto). Aplica a espigas, adaptadores, bridas y prearmadas.
cola: tipo de cola para espigas, bridas y prearmadas — "R12" si dice larga/R12, "INTERLOCK" si dice interlock/R13/R15, "" si no especifica (default R2/corta). Vacío para otros productos.
doble_hex: true solo si el usuario pide explícitamente "doble hexágono" o "c/hex". Default false.
ferrula_tm: solo para ferrulas — "si" por defecto (T/M tipo manulli); "no" SOLO si el usuario dice explícitamente "lisa" o "00210".
es_saludo: true si el mensaje es un saludo o consulta no relacionada con productos
Para ferrulas: el tipo debe incluir el subtipo SAE (ej: "FERRULA R1", "FERRULA R2", "FERRULA R12"). Aliases: 1sn/2sn/at = R2, 4SH/4SP = R12, R13/R15/interlock = R13-R15, t/m/manulli/tipo manulli = ferrula_tm="si" (default). lisa/00210 = ferrula_tm="no".
Medidas nominales (código 2 dígitos pegado al tipo → pulgadas): 02→1/8 | 03→3/16 | 04→1/4 | 05→5/16 | 06→3/8 | 08→1/2 | 10→5/8 | 12→3/4 | 14→7/8 | 16→1 | 20→1 1/4 | 24→1 1/2 | 32→2 — Ej: "JIC16"=1", "NPT08"=1/2". IMPORTANTE: NO redondees medidas al tamaño más cercano — si el usuario pide 3/16", el campo medida debe ser "3/16", no "1/4".
Dos tipos de rosca distintos en un pedido (NPT+JIC, BSP+ORFS, etc.) → ADAPTADOR: tipo="ADAP MACHO X1 X HEMBRA X2". Mismo tipo → ESPIGA con medidas=[terminal, espiga].
Aliases de marcas: JDE=JDEFLEX, VITI=VITILLO, MACTU=MACTUBI
Aliases de tipos: casco/casquillo = FERRULA | gir/girat = GIRATORIO | hex = HEXAGONAL | red/reductor = REDUCTOR | forx/orx = ORFS | bssp = BSP (typo frecuente) | bspp = BSPP | bspt = BSPT | silicona recta = MANG SILICONA RECTA | silicona codo 90 = MANG SILICONA CODO 90 | silicona codo 45 = MANG SILICONA CODO 45 | silicona codo = MANG SILICONA CODO | silicona corrugada = MANG SILICONA CORRUGADA | silicona radiador = MANG SILICONA RADIADOR | silicona = MANG SILICONA | pu/poliuretano = MANG PU | tapon macho milimetrico = TAPON MACHO METRICO
ESPIGA NPT sin indicar macho/hembra → default MACHO: tipo="ESPIGA MACHO NPT". Solo para NPT; otros tipos mantienen HEMBRA por defecto.
Si es_saludo es true, deja todos los demás campos vacíos.
Para el campo tipo: elige el grupo más general que aplique — si el usuario no especificó subtipo (R1/R2/R12/etc.), usa el prefijo base (ej: "ESPIGA MACHO NPT" en lugar de "ESPIGA MACHO NPT R2").
IMPORTANTE: Analiza SOLO el mensaje actual del usuario. Ignora las respuestas previas del asistente."""

# ── Prompt 2: asistente conversacional ───────────────────────────────────────
_AGENT_PROMPT = """\
Eres Lutong, el asistente comercial de CISGE, distribuidora peruana de mangueras hidráulicas y accesorios. Atiendes vendedores por WhatsApp.
PERSONALIDAD: Profesional, directo, respuestas cortas adaptadas a WhatsApp.
PRODUCTOS: Mangueras hidráulicas (R1,R2,4SH,4SP), aire, succión, espigas, ferrulas, adaptadores, válvulas. Marcas: QF, JDEFLEX, VITILLO, MACTUBI, AF. Precios en dólares + IGV.
INFORMACIÓN DE LA EMPRESA:
Razón social: Comercial CISGE S.A.C.
RUC: 20511843783
Sede principal: Jr. Edmundo Moreau 931, Lima (Lima Centro)
Sede San Luis: Calle Río Piura 120, San Luis
Teléfonos: (01) 451-0788 / (01) 452-5052
Web: www.comercialcisgesac.com.pe
Fundada: 2008
Rubro: Importación y distribución de mangueras hidráulicas, conexiones, adaptadores y accesorios industriales.
Horario de atención: Lunes a viernes de 8:00 a.m. a 5:30 p.m.

REGLAS:
- Tu nombre es Lutong. Si te preguntan quién eres o cómo te llamas, responde que eres Lutong, el asistente de CISGE.
- Si preguntan por información de la empresa no listada aquí, indica que no tienes ese dato y sugiere llamar al (01) 451-0788.
- Nunca inventes datos de la empresa.
- Si el usuario saluda o hace una consulta no relacionada con productos, responde cordialmente y pregunta en qué puedes ayudarle. No pidas tipo/medida/marca ante un saludo.
- Nunca inventes precios ni stock, solo usa las tools.
- Respuestas en español, formato WhatsApp (sin markdown).
- Cuando tengas suficiente información para buscar un producto, llama la tool directamente sin preguntar de nuevo.
- Cuando el usuario use pronombres o referencias vagas como 'el r1', 'ese producto', 'en colonial', siempre revisa el historial de la conversación para identificar a qué producto se refiere antes de pedir aclaraciones.
- Cuando el usuario envíe múltiples productos en un solo mensaje separados por comas o saltos de línea con cantidades, interpreta cada línea como un ítem de cotización y llama a tool_generar_cotizacion con todos los ítems juntos sin pedir confirmación previa.
- Cuando el usuario pregunte 'en qué marcas tienes X', 'qué marcas tienen X' o similar: busca el producto con tool_buscar_producto, extrae las marcas únicas de los resultados y lista SOLO las marcas disponibles con sus precios aproximados. No listes todos los productos."""

_TOOL_MAP = {
    "tool_buscar_producto":    tool_buscar_producto,
    "tool_ver_stock":          tool_ver_stock,
    "tool_generar_cotizacion": tool_generar_cotizacion,
}


async def agente_cisge(mensaje: str, numero_wa: str) -> str:
    """E1 código exacto → E2 prefijo → GPT parser → E3 campos → E4 GPT."""
    # ── Pre-check: intención de marcas → GPT directo ──────────────────────────
    if _INTENT_MARCAS.search(mensaje):
        logger.info(f"agente [{numero_wa}]: pregunta de marcas → GPT directo")
        respuesta = await _gpt_conversacional(mensaje, numero_wa)
        guardar_mensajes(numero_wa, mensaje, respuesta)
        return respuesta

    # ── Pre-check: ficha técnica ──────────────────────────────────────────────
    if _INTENT_FICHA.search(mensaje):
        # Extraer prefijo de grupo: buscar código en el mensaje, tomar primeros 3 chars
        texto_ft = re.sub(r'\b(ficha\s+t[eé]cnica|ficha\s+tecnica|ficha)\b', '', mensaje, flags=re.IGNORECASE).strip()
        texto_ft = re.sub(r'\b(del?|de\s+la|para\s+el?|el|la)\b', '', texto_ft, flags=re.IGNORECASE).strip()
        grupo = ""
        if texto_ft:
            r_ft = buscar_por_codigo(texto_ft)
            if r_ft.empty:
                r_ft = buscar_por_codigo_prefijo(texto_ft)
            if not r_ft.empty:
                grupo = r_ft.iloc[0]["codigo"][:3]
        # Fallback: primer bloque de 3 dígitos en el mensaje original
        if not grupo:
            m_g = re.search(r'\b(\d{3})', mensaje)
            if m_g:
                grupo = m_g.group(1)
        if grupo:
            logger.info(f"agente [{numero_wa}]: ficha técnica → grupo {grupo}")
            try:
                enviado = await _enviar_ficha_wa(numero_wa, grupo)
                if enviado:
                    guardar_mensajes(numero_wa, mensaje, f"[ficha técnica grupo {grupo} enviada]")
                    return ""
            except Exception as e:
                logger.error(f"agente [{numero_wa}]: error enviando ficha: {e}")
            respuesta = f"No tengo ficha técnica cargada para el grupo {grupo} aún."
        else:
            respuesta = "¿De qué producto necesitas la ficha técnica? Indícame el código o nombre."
        guardar_mensajes(numero_wa, mensaje, respuesta)
        return respuesta

    # Limpiar cantidad y descuento para búsquedas de código
    cantidad  = extraer_cantidad(mensaje)
    descuento = extraer_descuento(mensaje)
    texto_cod = re.sub(r'\bx\s*\d+(?!\s*/\d)', '', mensaje).strip()
    texto_cod = re.sub(r'\b\d+\s*%', '', texto_cod).strip()

    # ── E1: código exacto ─────────────────────────────────────────────────────
    exactos = buscar_por_codigo(texto_cod)
    if not exactos.empty:
        logger.info(f"agente [{numero_wa}]: E1 código exacto ({exactos.iloc[0]['codigo']})")
        respuesta = (formatear_resultado(exactos.iloc[0], cantidad, descuento)
                     if len(exactos) == 1 else formatear_multi_marca(exactos))
        guardar_mensajes(numero_wa, mensaje, respuesta)
        return respuesta

    # ── E2: prefijo de código ─────────────────────────────────────────────────
    if len(texto_cod) >= 3:
        prefijo = buscar_por_codigo_prefijo(texto_cod)
        # Fallback VITILLO: letra de color extra (VT-TH1SNN08 → VT-TH1SN08)
        if prefijo.empty:
            m_vt = re.match(r'^(.+)([A-Z])(\d{2}(?:SL)?)$', texto_cod.upper().strip())
            if m_vt:
                alt = m_vt.group(1) + m_vt.group(3)
                alt_r = buscar_por_codigo_prefijo(alt)
                if not alt_r.empty:
                    prefijo = alt_r
                    logger.info(f"agente [{numero_wa}]: E2 VITILLO corregido → {alt}")

        # Fallback: si el texto tiene espacios, probar tokens numéricos como prefijo
        # Ej: "Ferrula 03310" → probar "03310" | "021-06-06 dme" → probar "021-06-06" + filtrar por DME
        if prefijo.empty and ' ' in texto_cod:
            tokens = texto_cod.split()
            for token in tokens:
                if token[0].isdigit() and len(token) >= 3:
                    alt_tok = buscar_por_codigo_prefijo(token)
                    if not alt_tok.empty:
                        prefijo = alt_tok
                        logger.info(f"agente [{numero_wa}]: E2 token numérico → {token}")
                        # Si hay tokens no-numéricos, buscar si alguno es una marca
                        marcas_cat = set(motor_df["marca"].str.upper().dropna().unique())
                        for otro in tokens:
                            if otro == token:
                                continue
                            otro_up = otro.upper()
                            if otro_up in marcas_cat:
                                f = prefijo[prefijo["marca"].str.upper() == otro_up]
                                if not f.empty:
                                    prefijo = f
                                    logger.info(f"agente [{numero_wa}]: E2 marca → {otro_up}")
                                    break
                            elif otro.lower() in _aliases_marcas:
                                marca_real = _aliases_marcas[otro.lower()]
                                f = prefijo[prefijo["marca"].str.upper() == marca_real]
                                if not f.empty:
                                    prefijo = f
                                    logger.info(f"agente [{numero_wa}]: E2 alias marca {otro} → {marca_real}")
                                    break
                        break

        if len(prefijo) == 1:
            fila = prefijo.iloc[0]
            logger.info(f"agente [{numero_wa}]: E2 prefijo único ({fila['codigo']})")
            respuesta = formatear_resultado(fila, cantidad, descuento)
            guardar_mensajes(numero_wa, mensaje, respuesta)
            return respuesta
        elif 1 < len(prefijo) <= 15:
            logger.info(f"agente [{numero_wa}]: E2 prefijo {len(prefijo)} resultados")
            respuesta = (formatear_multi_marca(prefijo) if prefijo["codigo"].nunique() == 1
                         else formatear_lista(prefijo, f"Encontré {len(prefijo)} productos con ese prefijo:"))
            guardar_mensajes(numero_wa, mensaje, respuesta)
            return respuesta
        elif len(prefijo) > 15:
            logger.info(f"agente [{numero_wa}]: E2 prefijo {len(prefijo)} resultados → lista truncada")
            lista = formatear_lista(prefijo.head(15), f"Mostrando 15 de {len(prefijo)} productos con ese prefijo:")
            respuesta = lista + "\n_Hay más resultados — escribe más caracteres para filtrar._"
            guardar_mensajes(numero_wa, mensaje, respuesta)
            return respuesta

    # ── GPT parser: lenguaje natural → campos estructurados ───────────────────
    parsed = await _parsear(mensaje)

    # Post-proceso: si el parser no extrajo marca, escanear el mensaje contra _aliases_marcas
    if not parsed.get("marca") and _aliases_marcas:
        msg_lo = mensaje.lower()
        for alias in sorted(_aliases_marcas, key=len, reverse=True):
            if re.search(rf'\b{re.escape(alias)}\b', msg_lo):
                parsed["marca"] = _aliases_marcas[alias]
                break

    logger.info(f"agente [{numero_wa}]: parser → {parsed}")

    if parsed.get("es_saludo"):
        respuesta = await _gpt_conversacional(mensaje, numero_wa)
        guardar_mensajes(numero_wa, mensaje, respuesta)
        return respuesta

    # ── E3: buscar_por_tipo_medida_marca con campos del parser ────────────────
    respuesta_e3 = _buscar_con_parsed(parsed)
    if respuesta_e3:
        logger.info(f"agente [{numero_wa}]: E3 encontró resultado")
        guardar_mensajes(numero_wa, mensaje, respuesta_e3)
        return respuesta_e3

    # ── E4: GPT conversacional con tools ──────────────────────────────────────
    logger.info(f"agente [{numero_wa}]: E4 GPT conversacional")
    respuesta = await _gpt_conversacional(mensaje, numero_wa)
    guardar_mensajes(numero_wa, mensaje, respuesta)
    return respuesta


def _grupos_disponibles() -> str:
    grupos = sorted(motor_df["grupo"].dropna().unique().tolist())
    return ", ".join(f'"{g}"' for g in grupos)


def _subfamilias_disponibles() -> str:
    subs = sorted(motor_df["subfamilia"].dropna().unique().tolist())
    return ", ".join(f'"{s}"' for s in subs)


_SUBFAMILIAS_VALIDAS: set = set()


def _validar_subfamilias(subfamilias: list) -> list | None:
    global _SUBFAMILIAS_VALIDAS
    if not _SUBFAMILIAS_VALIDAS:
        _SUBFAMILIAS_VALIDAS = set(motor_df["subfamilia"].dropna().unique())
    validas = [s for s in (subfamilias or []) if s in _SUBFAMILIAS_VALIDAS]
    return validas if validas else None


async def _parsear(mensaje: str) -> dict:
    """Parser interno: lenguaje natural → JSON estructurado. Sin tools."""
    try:
        prompt = (
            _PARSER_PROMPT
            + f"\n\nSubfamilias válidas (usa SOLO estas en el campo 'subfamilias'):\n{_subfamilias_disponibles()}"
        )
        messages = [{"role": "system", "content": prompt},
                    {"role": "user", "content": mensaje}]
        completion = await _client.chat.completions.create(
            model=_MODEL, messages=messages, temperature=0,
        )
        return json.loads(completion.choices[0].message.content.strip())
    except Exception as e:
        logger.warning(f"_parsear error: {e}")
        return {"es_saludo": False}


# Prioridad de marca por subfamilia para respuestas de imagen
_MARCA_PRIORIDAD_IMAGEN: dict[str, list[str]] = {
    "ESPIGAS I":      ["LT", "XY"],
    "ESPIGAS II":     ["LT", "XY"],
    "BRIDAS":         ["XY", "LT"],
    "FERRULAS":       ["LT", "XY"],
    "ADAPTADORES I":  ["DME"],
    "ADAPTADORES II": ["DME"],
}


def _elegir_fila_por_prioridad(resultados, cantidad: int):
    """Elige una sola fila según prioridad de marca y stock suficiente."""
    from app.motor import _stock_total
    subfamilia = resultados.iloc[0]["subfamilia"]
    prioridad  = _MARCA_PRIORIDAD_IMAGEN.get(subfamilia, [])

    # 1. Marca prioritaria con stock >= cantidad pedida
    for marca in prioridad:
        filas = resultados[resultados["marca"].str.upper() == marca.upper()]
        if not filas.empty and _stock_total(filas.iloc[0]) >= cantidad:
            return filas.iloc[0]
    # 2. Marca prioritaria sin importar stock
    for marca in prioridad:
        filas = resultados[resultados["marca"].str.upper() == marca.upper()]
        if not filas.empty:
            return filas.iloc[0]
    # 3. Cualquier marca con mayor stock
    stocks = resultados.apply(_stock_total, axis=1)
    return resultados.loc[stocks.idxmax()]


def _buscar_fila_imagen(parsed: dict, cantidad: int):
    """Busca y selecciona la fila óptima para un ítem de imagen. Devuelve pd.Series o None."""
    # E1: si GPT extrajo un código de catálogo explícito, buscarlo primero
    codigo = (parsed.get("codigo") or "").strip()
    if codigo:
        fila_e1 = buscar_por_codigo(codigo)
        if not fila_e1.empty:
            return fila_e1.iloc[0]

    tipo       = parsed.get("tipo") or None
    medida     = parsed.get("medida") or None
    medidas    = parsed.get("medidas") or []
    marca      = parsed.get("marca") or None
    presion    = parsed.get("presion") or None
    angulo     = parsed.get("angulo") or None
    cola       = parsed.get("cola") or None
    doble_hex  = bool(parsed.get("doble_hex"))
    ferrula_tm = parsed.get("ferrula_tm") or ""
    subfamilias = _validar_subfamilias(parsed.get("subfamilias") or [])

    if not any([tipo, medida, marca, subfamilias]):
        return None

    resultados = buscar_por_tipo_medida_marca(
        tipo=tipo, medida=medida, marca=marca, presion=presion,
        subfamilias=subfamilias, medidas=medidas or None,
        angulo=angulo, cola=cola, doble_hex=doble_hex, ferrula_tm=ferrula_tm,
    )
    if resultados.empty:
        return None
    if len(resultados) == 1:
        return resultados.iloc[0]
    subfamilia = resultados.iloc[0]["subfamilia"]
    if subfamilia in _MARCA_PRIORIDAD_IMAGEN:
        return _elegir_fila_por_prioridad(resultados, cantidad)
    stocks = resultados.apply(_stock_total, axis=1)
    return resultados.loc[stocks.idxmax()]


def _buscar_con_parsed(parsed: dict, imagen_cantidad: int | None = None) -> str:
    """E3: llama a buscar_por_tipo_medida_marca() con campos del parser."""
    tipo        = parsed.get("tipo") or None
    medida      = parsed.get("medida") or None
    # Normalizar medida mm: "19mm" → "19" (por si GPT incluye la unidad)
    if medida:
        _mm = re.match(r'^(\d+)\s*mm$', medida.strip(), re.IGNORECASE)
        if _mm:
            medida = _mm.group(1)
    medidas     = parsed.get("medidas") or []
    marca       = parsed.get("marca") or None
    presion     = parsed.get("presion") or None
    color       = parsed.get("color") or None
    angulo      = parsed.get("angulo") or None
    cola        = parsed.get("cola") or None
    doble_hex   = bool(parsed.get("doble_hex"))
    ferrula_tm  = parsed.get("ferrula_tm") or ""
    subfamilias_raw = parsed.get("subfamilias") or []
    subfamilias = _validar_subfamilias(subfamilias_raw)
    if subfamilias_raw and subfamilias_raw != (subfamilias or []):
        logger.info(f"E3: subfamilias filtradas: {subfamilias_raw} → {subfamilias}")

    if not any([tipo, medida, marca, subfamilias]):
        logger.info("E3: sin campos suficientes para buscar")
        return ""

    subtipo_ht = color if tipo == "HT" else None
    logger.info(
        f"E3: buscar_por_tipo_medida_marca("
        f"tipo={tipo!r}, medida={medida!r}, medidas={medidas}, marca={marca!r}, "
        f"presion={presion!r}, angulo={angulo!r}, cola={cola!r}, subfamilias={subfamilias})"
    )
    resultados = buscar_por_tipo_medida_marca(
        tipo=tipo, medida=medida, marca=marca, presion=presion,
        subtipo=subtipo_ht, subfamilias=subfamilias, medidas=medidas or None,
        angulo=angulo, cola=cola, doble_hex=doble_hex, ferrula_tm=ferrula_tm,
    )
    logger.info(f"E3: DataFrame → {len(resultados)} filas")

    if resultados.empty:
        return ""

    if len(resultados) == 1:
        cant = imagen_cantidad or 1
        return formatear_resultado(resultados.iloc[0], cant)

    # Modo imagen: subfamilias con prioridad de marca → mostrar solo 1 resultado
    if imagen_cantidad is not None:
        subfamilia = resultados.iloc[0]["subfamilia"]
        if subfamilia in _MARCA_PRIORIDAD_IMAGEN:
            fila = _elegir_fila_por_prioridad(resultados, imagen_cantidad)
            logger.info(f"E3 imagen: seleccionado {fila['codigo']} — {fila['marca']}")
            return formatear_resultado(fila, imagen_cantidad)

    if resultados["codigo"].nunique() == 1:
        return formatear_multi_marca(resultados)

    if len(resultados) <= 12:
        return formatear_lista(resultados, f"Encontré {len(resultados)} opciones:")

    marcas = resultados["marca"].unique()
    meds   = resultados["medida_cod"].dropna().unique()[:8]
    detalle = []
    if not marca:
        detalle.append(f"Marca: {', '.join(marcas)}")
    if not medida:
        detalle.append(f"Medida: {', '.join(meds)}")
    return (
        f"Encontré {len(resultados)} productos.\n\n"
        "¿Puedes especificar?\n" + "\n".join(detalle)
    )


async def _gpt_conversacional(mensaje: str, numero_wa: str) -> str:
    """E4: GPT con tools para saludos y consultas que E3 no pudo resolver."""
    historial = cargar_historial(numero_wa)
    messages = [{"role": "system", "content": _AGENT_PROMPT}]
    messages.extend(historial)
    messages.append({"role": "user", "content": mensaje})
    return await _llamar_gpt(messages)


async def _llamar_gpt(messages: list) -> str:
    """Llama a GPT con tools. Resuelve hasta 5 rondas de tool calls."""
    for _ in range(5):
        completion = await _client.chat.completions.create(
            model=_MODEL, messages=messages, tools=TOOLS_SCHEMA, tool_choice="auto",
        )
        msg = completion.choices[0].message

        if not msg.tool_calls:
            return msg.content or ""

        messages.append(msg)
        for tc in msg.tool_calls:
            nombre = tc.function.name
            try:
                args = json.loads(tc.function.arguments)
            except json.JSONDecodeError:
                args = {}
            fn = _TOOL_MAP.get(nombre)
            if fn:
                try:
                    resultado = fn(**args)
                except Exception as e:
                    resultado = f"Error al ejecutar {nombre}: {e}"
                    logger.warning(resultado)
            else:
                resultado = f"Tool desconocida: {nombre}"
            messages.append({"role": "tool", "tool_call_id": tc.id, "content": resultado})

    completion = await _client.chat.completions.create(model=_MODEL, messages=messages)
    return completion.choices[0].message.content or ""


# ── Procesamiento de imágenes WhatsApp ────────────────────────────────────────

_WA_TOKEN = os.getenv("WHATSAPP_TOKEN")
_PHONE_ID = os.getenv("PHONE_NUMBER_ID")

_OCR_PROMPT = """\
Eres un extractor de texto de imágenes para CISGE, distribuidora industrial peruana.
Extrae el texto de la imagen línea por línea, tal como aparece escrito.
IMPORTANTE: Debes extraer ABSOLUTAMENTE TODAS las líneas del pedido, de arriba a abajo, sin omitir ninguna aunque la escritura sea difícil de leer. No te detengas hasta haber procesado hasta la última línea visible.
Para palabras de escritura ambigua o difícil de leer, elige el término más probable del vocabulario del dominio.
Si la imagen no contiene texto legible, responde exactamente: SIN_TEXTO

Vocabulario frecuente (úsalo para resolver ambigüedad en escritura a mano):
  casco, ferrula, espiga, adaptador, codo, niple, union, valvula, manguera, rollo, liso, larga, reduccion, tapon, brida, prearmada.
Correcciones OCR conocidas: "benda"/"brenda"/"bnda" → "brida" | "ferula"/"ferulla" → "ferrula" | "espinga"/"espgia" → "espiga".
Ejemplo: una palabra que podría leerse "casco" o "codo" — si va seguida de R1/R2/R12, es "casco" (los cascos llevan subtipo SAE, los codos no)."""


async def procesar_imagen_whatsapp(image_id: str, numero_wa: str) -> str:
    """Descarga imagen de WhatsApp, extrae texto con GPT-4o Vision y procesa cada línea."""
    headers_wa = {"Authorization": f"Bearer {_WA_TOKEN}"}

    # Paso 1 — obtener URL de la imagen desde Meta
    async with httpx.AsyncClient() as client:
        r = await client.get(
            f"https://graph.facebook.com/v19.0/{image_id}",
            headers=headers_wa, timeout=10,
        )
        r.raise_for_status()
        image_url = r.json().get("url")
        if not image_url:
            return "No pude obtener la imagen. Intenta de nuevo."

        # Paso 2 — descargar imagen en memoria
        r2 = await client.get(image_url, headers=headers_wa, timeout=20)
        r2.raise_for_status()
        image_b64 = base64.b64encode(r2.content).decode()
        mime = r2.headers.get("content-type", "image/jpeg").split(";")[0]

    # Paso 3 — OCR con GPT-4o Vision (texto plano, sin JSON)
    try:
        completion = await _client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": _OCR_PROMPT},
                {"role": "user", "content": [
                    {"type": "image_url", "image_url": {
                        "url": f"data:{mime};base64,{image_b64}"
                    }}
                ]},
            ],
            max_tokens=4096,
            temperature=0,
        )
        texto = completion.choices[0].message.content.strip()
        logger.info(f"OCR [{numero_wa}]: texto extraído ({len(texto)} chars): {texto[:200]!r}")
    except Exception as e:
        logger.warning(f"OCR error [{numero_wa}]: {type(e).__name__}: {e}")
        return "Hubo un problema leyendo la imagen. Intenta de nuevo o escribe la lista directamente."

    # Paso 4 — imagen sin texto legible
    if not texto or texto == "SIN_TEXTO":
        return "No pude leer texto en la imagen. ¿Puedes enviar una foto más clara o escribir la lista directamente?"

    # Paso 5 — GPT parsea el texto OCR a campos estructurados (tipo/medida/marca/cantidad)
    texto = _corregir_codo_ocr(texto)
    n_brutas = len([l for l in texto.splitlines() if l.strip()])
    logger.info(f"OCR [{numero_wa}]: {n_brutas} líneas brutas extraídas")
    try:
        parsed_list = await _parsear_lineas_imagen(texto)
        parsed_list = _enriquecer_tipo_ferrula(parsed_list)
        parsed_list = _corregir_medidas_ocr(parsed_list)
        logger.info(f"OCR [{numero_wa}]: {len(parsed_list)} ítems parseados")
    except Exception as e:
        logger.warning(f"_parsear_lineas_imagen error: {e}")
        guardar_mensajes(numero_wa, "[imagen]", "No pude interpretar la lista. Escríbela directamente.")
        return "No pude interpretar la lista. Escríbela directamente."

    # Paso 6 — E3 local para cada ítem; construir filas Excel con todo en orden
    rows_excel = []
    for parsed in parsed_list:
        linea_orig = parsed.get("linea_original", "?")
        cantidad   = int(parsed.get("cantidad", 1))
        fila_motor = None
        try:
            if _buscar_con_parsed(parsed, imagen_cantidad=cantidad):
                fila_motor = _buscar_fila_imagen(parsed, cantidad)
        except Exception as exc:
            logger.warning(f"Excel capture error '{linea_orig}': {exc}")

        if fila_motor is not None:
            precio_u = float(fila_motor["precio"])
            subtotal = round(precio_u * cantidad, 2)
            rows_excel.append({
                "linea_original": linea_orig,
                "codigo":         fila_motor["codigo"],
                "descripcion":    fila_motor["descripcion"].title(),
                "cantidad":       cantidad,
                "precio_unit":    precio_u,
                "subtotal":       subtotal,
                "igv":            round(subtotal * IGV, 2),
                "total":          round(subtotal * (1 + IGV), 2),
                "encontrado":     True,
            })
        else:
            tipo_p   = parsed.get("tipo") or ""
            medida_p = parsed.get("medida") or (parsed.get("medidas") or [""])[0]
            if tipo_p and _tipo_existe_en_catalogo(tipo_p):
                nota   = "Tamaño no disponible en catálogo"
                motivo = "tamaño"
            elif tipo_p:
                nota   = "Producto no existe en el catálogo"
                motivo = "producto"
            else:
                nota   = "No reconocido en catálogo"
                motivo = "producto"
            if tipo_p:
                _registrar_demanda(tipo_p, medida_p, cantidad, linea_orig, numero_wa, motivo)
            rows_excel.append({"linea_original": linea_orig, "nota": nota, "encontrado": False})

    # Paso 7 — enviar solo Excel (texto OCR completo en orden: encontrados + no encontrados)
    guardar_mensajes(numero_wa, f"[imagen: {len(parsed_list)} ítems]", "Excel enviado")
    if not rows_excel:
        return "No encontré ningún producto en la imagen."
    try:
        excel_bytes = generar_excel_bytes(rows_excel)
        media_id    = await _subir_media_wa(excel_bytes, "cotizacion_cisge.xlsx")
        await _enviar_doc_wa(numero_wa, media_id, "cotizacion_cisge.xlsx")
        logger.info(f"OCR [{numero_wa}]: Excel enviado ({len(rows_excel)} filas)")
        return ""
    except Exception as exc:
        logger.warning(f"OCR [{numero_wa}]: Excel send error: {exc}")
        return "No pude enviar el Excel. Intenta de nuevo."


_PARSEAR_LINEAS_PROMPT = f"""\
Eres un parser para CISGE, distribuidora peruana de mangueras hidráulicas.
Recibirás texto OCR de una lista de productos. Para cada ítem con cantidad, extrae campos estructurados.

Vocabulario cerrado — usa EXACTAMENTE estos términos, sin expandir a nombre completo ni sinónimos:
Tipos de rosca: JIC | ORFS | BSP | BSPP | BSPT | NPT | SAE | METRIC | LIVIANA | PESADA | KOMATSU
Familias: ESPIGA | FERRULA | ADAPTADOR | MANGUERA | NIPLE | VALVULA | BRIDA | CASCO | REDUCCION | TAPON
Si el OCR entregó una palabra incierta, elige el término más cercano de la lista y úsalo tal cual.

Aliases (sinónimos, abreviaciones y errores OCR conocidos):
ores/o-rings/o-ring = ORFS | casco/casq/casquillo = FERRULA | gir/girat = GIRATORIO | hex = HEXAGONAL | red = REDUCTOR
bssp/bsps/bbsp = BSP
luvana/luvata/luvat/luvani = MM LIVIANA | pessao/pesao/pessoni/pesoni = MM PESADA
C-61/C61/cod61/code61/c-61 = CODE 61 | C-62/C62/cod62/code62/c-62 = CODE 62 | Cat/cat/CAT = CAT
benda/brend/bnda = BRIDA
tapon macho milimetrico = TAPON MACHO METRICO

Bridas (tipo = "BRIDA CODE 61", "BRIDA CODE 62" o "BRIDA CAT"):
- Si el texto contiene BRIDA o alias (benda/bnda) como familia, tipo = "BRIDA CODE 61" / "BRIDA CAT" etc. NUNCA "ESPIGA HEMBRA BRIDA ...".
- CODE 61, CODE 62, CAT van SOLO en el tipo, NUNCA en marca. El campo marca queda vacío para bridas.
- El prefijo "ESP." antes de BRIDA es ruido del vendedor — ignorarlo. "H" o "Hembra" antes de BRIDA también se ignoran.
  Ej: "ESP. BRIDA C-61 3/4x3/4" → tipo="BRIDA CODE 61", medida="3/4", marca=""
  Ej: "ESP. H Benda C-61 3/4x3/4" → tipo="BRIDA CODE 61", medida="3/4", marca=""
- Notación NxM (ej: 1x1, 3/4x3/4, 1x1/2): medida = segundo número (tamaño de manguera). Ignorar el primero.
- Si aparece un valor en mm intermedio (ej: 125mm, 150mm): ignorarlo — no va en ningún campo.
- angulo: extraer normalmente si aparece 90° o 45°.
- cola: "R12" si dice "large" o R12, "" si no especifica (R2 por defecto).

Espigas KOMATSU (tipo = "ESPIGA HEMBRA KOMATSU"):
- Formato MNNxMEDIDA (ej: M24x5/8, M18x1/2): medida = parte DESPUÉS del x (tamaño de manguera: "5/8", "1/2"). El MNN es el hilo del puerto Komatsu — ignorarlo.
- angulo: extraer normalmente ("90" si dice 90°, "45" si dice 45°, "" si no especifica = recta).
- doble_hex: true si dice "doble hex", "doble hexagono" o "DH". Por defecto false (estándar sin D).
- cola: "" por defecto (R2).

Espigas métricas (LIVIANA = métrica liviana, PESADA = métrica pesada — son tipos de rosca, NO marcas):
- tipo debe incluir "MM": "ESPIGA HEMBRA MM LIVIANA", "ESPIGA MACHO MM LIVIANA", "ESPIGA HEMBRA MM PESADA", etc.
- medida: extraer solo el diámetro métrico principal en formato "M12", "M14", "M18", etc. Ignorar pitch (x1.5) y tamaño de manguera al final.
- angulo: si hay 90 o 45 al final, extraerlo normalmente.
- NO poner M12/M14 en marca.
Para ferrulas: el tipo debe incluir el subtipo SAE si aparece (ej: "FERRULA R1", "FERRULA R2", "FERRULA R12"). No dejar solo "FERRULA" si hay un subtipo en la línea.
Aliases de subtipo ferrula: 4SH/4SP/4sh/4sp = R12 | 1SN/2SN/AT = R2 | R13/R15/INTERLOCK = R13-R15.
4SH y 4SP NO van en marca — son el subtipo de la ferrula. Ej: "FERRULA 2\" 4SH" → tipo="FERRULA R12", medida="2", marca="".
ferrula_tm: solo para ferrulas — "si" por defecto (T/M tipo manulli); "no" SOLO si dice explícitamente "lisa" o "00210".

Subfamilias válidas: "ESPIGAS I", "ESPIGAS II", "ADAPTADORES I", "ADAPTADORES II",
"FERRULAS", "MANGUERAS HIDRAULICAS", "MANGUERAS INDUSTRIALES", "VALVULAS",
"NIPLES", "CAMLOCK", "BRIDAS", "TUBERIAS HIDRAULICAS", "PREARMADAS", "MANOMETROS"

Devuelve SOLO un objeto JSON con campo "items" que contiene el array. Cada elemento:
{{"linea_original":"texto como aparece","codigo":"","tipo":"ESPIGA HEMBRA ORFS","medida":"1","medidas":[],"marca":"","cantidad":50,"angulo":"","cola":"","ferrula_tm":"si","subfamilias":["ESPIGAS I","ESPIGAS II"]}}
codigo: si la línea contiene explícitamente un código de catálogo CISGE (alfanumérico con guiones, ej: DCA-10L, QF-R1-1/2, OR-SHH-19, 26711-12-12), extraerlo aquí tal cual. Si no hay código explícito, dejar vacío.
Formato: {{"items": [{{...}}, {{...}}]}}

TABLA DE MEDIDAS NOMINALES (código 2 dígitos pegado al tipo → pulgadas):
03→3/16 | 04→1/4 | 05→5/16 | 06→3/8 | 08→1/2 | 10→5/8 | 12→3/4 | 14→7/8 | 16→1 | 20→1 1/4 | 24→1 1/2 | 32→2
Ej: "JIC16" = JIC 1", "NPT08" = NPT 1/2", "ORFS12" = ORFS 3/4".
IMPORTANTE: Respeta la medida EXACTA que aparece en el texto. Si dice "3/16", el campo medida/medidas debe ser "3/16". NO redondees al tamaño más cercano (ej: 3/16 ≠ 1/4). Si la medida no existe en catálogo, el sistema la detectará después.

ESTRUCTURA TERMINAL-ESPIGA (dos medidas/tipos separados por "-"):
Las espigas tienen dos lados: TERMINAL (manguera) y ESPIGA (rosca).
REGLA PRIMARIA: si la línea contiene las palabras "terminal" Y ("esp"/"espiga"), es SIEMPRE ESPIGA — nunca ADAPTADOR.
- Un solo tipo de rosca mencionado (o ninguno): es ESPIGA. El segundo lado hereda el mismo tipo; NO supongas JIC ni otro.
  Ej: "terminal JIC 16 - 12 Esp"  → tipo="ESPIGA HEMBRA JIC", medidas=["1","3/4"]
  Ej: "terminal NPT08 - 12 esp"   → tipo="ESPIGA MACHO NPT", medidas=["1/2","3/4"]  ← MACHO por defecto en NPT
  Ej: "terminal BSP12 - 08 esp"   → tipo="ESPIGA HEMBRA BSP", medidas=["3/4","1/2"]
  Ej: "JIC16-JIC12"               → tipo="ESPIGA HEMBRA JIC", medidas=["1","3/4"]
- Si NO se menciona ningún tipo de rosca, NO lo supongas: deja tipo="ESPIGA HEMBRA" sin especificar rosca.
  Ej: "terminal 08 - 12 esp"      → tipo="ESPIGA HEMBRA", medidas=["1/2","3/4"]
- ESPIGA NPT sin indicar macho/hembra → default MACHO (las espigas NPT son macho por defecto).
  Ej: "terminal NPT08 - 12 esp"   → tipo="ESPIGA MACHO NPT", medidas=["1/2","3/4"]  ← MACHO, no HEMBRA
- Distintos tipos de rosca AMBOS explícitos (y sin "terminal"+"esp"): es ADAPTADOR
  Ej: "NPT08 - JIC16 90"          → tipo="ADAP MACHO NPT X HEMBRA JIC", medidas=["1/2","1"], angulo="90"
  Ej: "BSP12-ORFS08"              → tipo="ADAP MACHO BSP X HEMBRA ORFS", medidas=["3/4","1/2"]
  Convención: primer tipo=MACHO, segundo tipo=HEMBRA. Número suelto al final (45/90)=angulo.

Reglas generales:
- "11/2" o "11/4" sin espacio → "1 1/2" / "1 1/4" en el campo medida
- Si una línea es encabezado sin cantidad (ej: "mang azul poliuretano"), úsala como contexto para las siguientes
- Tipo: usa el nombre más descriptivo posible (ej: "ESPIGA HEMBRA ORFS", no solo "ESPIGA")
- Medidas separadas por x (ej: "1/4x1/4", "3/8x1/2", "1/4x1/4x1/4x45°"): tomar SOLO las dos primeras como medidas=["N","M"], ignorar terceras en adelante. El ángulo al final (x45°, x90°) va en angulo, no en medidas. Dejar medida vacío. Excepción: bridas siguen su propia regla NxM.
- Si no hay cantidad explícita, usar 1
- angulo: "45" si dice 45°, "90" si dice 90°, "" si no especifica (recta)
- cola: "R12" si dice larga/R12, "INTERLOCK" si dice interlock/R13/R15, "" si no especifica (default R2)"""


async def _parsear_lineas_imagen(texto: str) -> list[dict]:
    """Una sola llamada GPT que convierte OCR a lista de dicts estructurados."""
    completion = await _client.chat.completions.create(
        model=_MODEL,
        messages=[
            {"role": "system", "content": _PARSEAR_LINEAS_PROMPT},
            {"role": "user",   "content": texto},
        ],
        temperature=0,
        response_format={"type": "json_object"},
    )
    raw = completion.choices[0].message.content.strip()
    resultado = json.loads(raw).get("items", [])
    return [l for l in resultado if isinstance(l, dict)]


# ── Excel ─────────────────────────────────────────────────────────────────────

def generar_excel_bytes(rows: list[dict]) -> bytes:
    """Genera un .xlsx en memoria. rows: lista ordenada de dicts con clave 'encontrado'.
    Filas reconocidas incluyen codigo/descripcion/precio; no reconocidas solo linea_original."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Cotizacion CISGE"

    headers = ["N°", "Línea Original", "Código CISGE", "Descripción", "Cantidad",
               "Precio Unit. USD", "Subtotal USD", "IGV (18%)", "Total USD"]
    ws.append(headers)

    fill_header  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_nf      = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # amarillo claro
    font_header  = Font(color="FFFFFF", bold=True)
    font_nf      = Font(italic=True, color="999999")
    center       = Alignment(horizontal="center")
    num_fmt      = '#,##0.00'

    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = center

    for n, item in enumerate(rows, start=1):
        row_num = ws.max_row + 1
        if item.get("encontrado"):
            ws.append([
                n,
                item["linea_original"],
                item["codigo"],
                item["descripcion"],
                item["cantidad"],
                item["precio_unit"],
                item["subtotal"],
                item["igv"],
                item["total"],
            ])
            ws.cell(row=row_num, column=1).alignment = center
            for col in range(6, 10):   # columnas F-I: precios
                ws.cell(row=row_num, column=col).number_format = num_fmt
        else:
            nota = item.get("nota", "No encontrado en catálogo")
            ws.append([n, item["linea_original"], "—", nota, None, None, None, None, None])
            for col in range(1, 10):
                ws.cell(row=row_num, column=col).fill = fill_nf
                ws.cell(row=row_num, column=col).font = font_nf
            ws.cell(row=row_num, column=1).alignment = center

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 12
    for col in ("F", "G", "H", "I"):
        ws.column_dimensions[col].width = 17

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _subir_media_wa(data: bytes, filename: str, mime: str = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet") -> str:
    """Sube un archivo al endpoint de media de Meta y devuelve el media_id."""
    url     = f"https://graph.facebook.com/v19.0/{_PHONE_ID}/media"
    headers = {"Authorization": f"Bearer {_WA_TOKEN}"}
    async with httpx.AsyncClient() as client:
        r = await client.post(
            url,
            headers=headers,
            files={"file": (filename, data, mime)},
            data={"messaging_product": "whatsapp"},
            timeout=30,
        )
        r.raise_for_status()
        media_id = r.json()["id"]
        logger.info(f"_subir_media_wa: subido {filename} → media_id={media_id}")
        return media_id


async def _enviar_doc_wa(numero: str, media_id: str, filename: str) -> None:
    """Envía un mensaje de tipo documento por WhatsApp."""
    url     = f"https://graph.facebook.com/v19.0/{_PHONE_ID}/messages"
    headers = {"Authorization": f"Bearer {_WA_TOKEN}", "Content-Type": "application/json"}
    body    = {
        "messaging_product": "whatsapp",
        "to":   numero,
        "type": "document",
        "document": {"id": media_id, "filename": filename},
    }
    async with httpx.AsyncClient() as client:
        r = await client.post(url, json=body, headers=headers, timeout=10)
        logger.info(f"_enviar_doc_wa: {r.status_code} → {numero}")
